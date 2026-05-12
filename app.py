import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy
import datetime
import re
import zipfile
import calendar
from openpyxl import load_workbook
from datetime import datetime

# ==========================================
# 1. INITIAL SETUP & THEME
# ==========================================

st.set_page_config(page_title="PAGASA RIDF Calculator", page_icon="⛈️", layout="wide")

# Modern PAGASA Style Interface using CSS
st.markdown("""
<style>

/* =========================
   MAIN APP
========================= */

.stApp {
    background:
        linear-gradient(
            135deg,
            #eef5ff 0%,
            #d8ebff 35%,
            #f7fbff 100%
        );
}

/* =========================
   REMOVE STREAMLIT PADDING
========================= */

.block-container {
    padding-top: 2rem;
    padding-left: 2rem;
    padding-right: 2rem;
    padding-bottom: 4rem;
    max-width: 1500px;
}

/* =========================
   HERO SECTION
========================= */

.hero-container {
    background:
        linear-gradient(
            135deg,
            #005BAC,
            #0077cc
        );

    border-radius: 28px;
    padding: 2.5rem;

    display: flex;
    align-items: center;
    gap: 2.5rem;

    box-shadow:
        0 10px 40px rgba(0,91,172,0.25);

    margin-bottom: 2rem;
}

.hero-logo {
    width: 130px;
    height: auto;
    filter: drop-shadow(0 5px 15px rgba(0,0,0,0.3));
    flex-shrink: 0;
}

.hero-right h1 {
    color: white !important;
    font-size: 2.8rem !important;
    margin: 0 !important;
    padding: 0 !important;
    font-weight: 800 !important;
    line-height: 1.2 !important;
    border: none !important; /* Removes any default bottom borders */
}

.hero-right p {
    color: rgba(255,255,255,0.9) !important;
    font-size: 1.2rem !important;
    margin-top: 0.5rem !important;
    margin-bottom: 0 !important;
    line-height: 1.4 !important;
}

/* =========================
   TABS
========================= */

.stTabs [data-baseweb="tab-list"] {
    gap: 1rem;
    background: rgba(255,255,255,0.55);
    padding: 0.5rem;
    border-radius: 18px;
}

.stTabs [data-baseweb="tab"] {
    height: 55px;
    padding-left: 25px;
    padding-right: 25px;
    border-radius: 14px;
    font-weight: 700;
    color: #005BAC;
    background: transparent;
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(
        135deg,
        #0077cc,
        #005BAC
    ) !important;

    color: white !important;
}

/* =========================
   SECTION CARDS
========================= */

.section-card {
    background: rgba(255,255,255,0.7);
    backdrop-filter: blur(10px);

    border-radius: 24px;

    padding: 2rem;

    margin-top: 1.5rem;
    margin-bottom: 2rem;

    border: 1px solid rgba(255,255,255,0.4);

    box-shadow:
        0 8px 30px rgba(0,0,0,0.08);
}

/* Change your .section-card selector to this */
div[data-testid="stVerticalBlock"]:has(> div.element-container > div.section-card-trigger) {
    background: rgba(255,255,255,0.7);
    backdrop-filter: blur(10px);
    border-radius: 24px;
    padding: 2.5rem;
    margin-top: 1.5rem;
    margin-bottom: 2rem;
    border: 1px solid rgba(255,255,255,0.4);
    box-shadow: 0 8px 30px rgba(0,0,0,0.08);
}       
/* =========================
   HEADINGS
========================= */

h1, h2, h3 {
    color: #003B73;
    font-weight: 800;
}

/* =========================
   BUTTONS
========================= */

.stButton > button {
    background:
        linear-gradient(
            135deg,
            #0077CC,
            #005BAC
        );

    color: white;

    border: none;
    border-radius: 16px;

    padding: 0.8rem 1.5rem;

    font-weight: 700;

    transition: 0.3s ease;

    box-shadow:
        0 6px 20px rgba(0,91,172,0.25);
}

.stButton > button:hover {
    transform: translateY(-3px);

    box-shadow:
        0 10px 25px rgba(0,91,172,0.35);
}

/* =========================
   FILE UPLOADER
========================= */

[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.75);

    border:
        2px dashed #4DA8FF;

    border-radius: 24px;

    padding: 2rem;

    box-shadow:
        inset 0 0 20px rgba(0,91,172,0.04);
}

/* =========================
   ALERTS
========================= */

.stSuccess,
.stWarning,
.stInfo,
.stError {
    border-radius: 18px !important;
}

/* =========================
   DATAFRAME
========================= */

[data-testid="stDataFrame"] {
    border-radius: 20px;
    overflow: hidden;
    border: 1px solid #dce8f5;
}

/* =========================
   SIDEBAR
========================= */

[data-testid="stSidebar"] {
    background:
        linear-gradient(
            180deg,
            #003B73,
            #00264d
        );
}

[data-testid="stSidebar"] * {
    color: white !important;
}

/* =========================
   METRICS
========================= */

.metric-box {
    background:
        linear-gradient(
            135deg,
            #0077CC,
            #005BAC
        );

    padding: 1.5rem;

    border-radius: 22px;

    color: white;

    box-shadow:
        0 10px 30px rgba(0,91,172,0.22);
}

.metric-value {
    font-size: 2rem;
    font-weight: 800;
}

.metric-label {
    opacity: 0.85;
}
/* =========================
   TRAIN STATION TRACKER
========================= */
.tracker-container {
    padding: 1.5rem 1rem;
    position: relative;
    margin-left: 10px;
}

/* The vertical track line */
.tracker-container::before {
    content: '';
    position: absolute;
    left: 19px;
    top: 2.5rem;
    bottom: 2.5rem;
    width: 2px;
    background: rgba(255, 255, 255, 0.2);
    z-index: 0;
}

.step-item {
    display: flex;
    align-items: center;
    gap: 15px;
    padding: 12px 0;
    position: relative;
    z-index: 1;
    color: rgba(255, 255, 255, 0.4);
}

/* The station circles */
.station-dot {
    width: 12px;
    height: 12px;
    border-radius: 50%;
    background: #1a3a5a;
    border: 2px solid rgba(255,255,255,0.3);
    flex-shrink: 0;
}

.step-active .station-dot {
    background: #4DA8FF;
    box-shadow: 0 0 10px #4DA8FF;
    border-color: white;
    transform: scale(1.3);
}

.step-complete .station-dot {
    background: #00ff88;
    border-color: #00ff88;
}

.step-active { color: white !important; font-weight: bold; }
.step-complete { color: #00ff88 !important; }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'uploader_key' not in st.session_state:
    st.session_state['uploader_key'] = 0

# --- PERSISTENT TRACKER LOGIC ---
steps = [
    {"label": "1: HMDAS-12 Batch Upload"},
    {"label": "2: Validated & Fixed"},
    {"label": "3: HMDAS-13 Upload"},
    {"label": "4: HMDAS-14 Built"},
    {"label": "5: Compilation"},
    {"label": "Annual Summary Complete"}
]

# Map your app logic to the step index
current_step = 0
if 'files' in locals() and files: current_step = 1
if st.session_state.get('ready_batch'): current_step = 2

if st.session_state.get('hmdas13_ready'): current_step = 3 
if st.session_state.get('h14_complete'): current_step = 4
if st.session_state.get('h15_complete'): current_step = 5

with st.sidebar:
    st.markdown("### RIDF Progress")
    st.markdown('<div class="tracker-container">', unsafe_allow_html=True)
    
    for i, step in enumerate(steps):
        status_class = "step-item"
        if i == current_step: status_class += " step-active"
        elif i < current_step: status_class += " step-complete"
        
        st.markdown(f"""
            <div class="{status_class}">
                <div class="station-dot"></div>
                <span>{step['label']}</span>
            </div>
        """, unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("---")
# --- 1. CORE LOGIC FUNCTIONS ---

def scan_files(uploaded_files):
    """Scans and returns the specific rows and filenames for the report."""
    error_log = {}
    for up_file in uploaded_files:
        wb = openpyxl.load_workbook(up_file, data_only=True)
        ws = wb.active
        rows_found = []
        
        # Scan Column B (Index 2) from Row 7 onwards
        for r in range(7, ws.max_row + 1):
            prev = ws.cell(row=r-1, column=2).value
            curr = ws.cell(row=r, column=2).value
            if isinstance(prev, (int, float)) and isinstance(curr, (int, float)):
                if prev > curr:
                    rows_found.append(r)

        if rows_found:
            error_log[up_file.name] = rows_found
    return error_log

def process_batch_for_aggregation(uploaded_files, error_log):
    """Fixes errors, restores formulas, and identifies files for the audit ZIP."""
    final_batch = []
    fixed_files_only = []

    for up_file in uploaded_files:
        wb = openpyxl.load_workbook(up_file)
        ws = wb.active
        is_modified = False

        if up_file.name in error_log:
            is_modified = True
            for r in error_log[up_file.name]:  # Use the specific rows from scan
                if r <= ws.max_row:
                    prev_b = ws.cell(row=r-1, column=2).value
                    curr_b = ws.cell(row=r, column=2).value
                    if isinstance(prev_b, (int, float)) and isinstance(curr_b, (int, float)):
                        if prev_b > curr_b:
                            ws.cell(row=r, column=2).value = prev_b

        out = io.BytesIO()
        wb.save(out)
        
        file_data = {"name": up_file.name, "data": out.getvalue()}
        final_batch.append(file_data)
        
        # Add to ZIP list only if it was an "Error" file 
        if is_modified:
            fixed_files_only.append(file_data)
            
    return final_batch, fixed_files_only

def reset_app():
    # 1. Clear flags used by the Sidebar Tracker
    st.session_state['ready_batch'] = None
    st.session_state['hmdas13_ready'] = False
    st.session_state['h14_complete'] = False
    st.session_state['h15_complete'] = False
    
    # 2. Increment the uploader key to force clear file widgets
    st.session_state['uploader_key'] += 1
    
    # 3. Force a rerun to refresh the UI immediately
    st.rerun()

# --- 2. INTERFACE & STAGE 1: BATCH VALIDATION ---\
from openpyxl import load_workbook


# Add this to your existing Tab structure

st.markdown(f"""
<div class="hero-container">
    <div class="hero-left">
        <img src="https://upload.wikimedia.org/wikipedia/commons/thumb/0/0b/Philippine_Atmospheric%2C_Geophysical_and_Astronomical_Services_Administration_%28PAGASA%29_logo.svg/1280px-Philippine_Atmospheric%2C_Geophysical_and_Astronomical_Services_Administration_%28PAGASA%29_logo.svg.png" class="hero-logo">
    </div>
    <div class="hero-right">
        <h1>RIDF Generator System</h1>
        <p>Rain Intensity Duration Frequency Processing Platform<br>for Hydrometeorological Analysis</p>
    </div>
</div>
""", unsafe_allow_html=True)



# Define the path to your internal HMDAS-15 template
HMDAS15_TEMPLATE_PATH = "templates/HMDAS-15 template.xlsm" 

# Create Tabs in your UI
tab1, tab2 = st.tabs(["Monthly (HMDAS-14)", "Annual (HMDAS-15)"])

with tab1:

    st.subheader("📂 Stage 1 · Batch Validation & Correction")

    if st.button("🧹 Clear & Reset Batch"):
        reset_app()

    files = st.file_uploader(
        "Upload Batch HMDAS-12 Files",
        type=['xlsx', 'xlsm'],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state['uploader_key']}"
    )

    if files:
        # Consistency Check: Station, Year, Month
        name_pattern = r"\s+(.+)\s+(\d{4})(\d{2})(\d{2})"
        batch_meta = []
        valid_batch = True

        for f in files:
            match = re.search(name_pattern, f.name)
            if not match:
                st.error(f"❌ **Naming Error:** `{f.name}` does not follow naming convention.")
                valid_batch = False
                break
            station, year, month, day = match.groups()
            batch_meta.append({'station': station.strip(), 'year': year, 'month': month, 'day': day, 'name': f.name})

        if valid_batch:
            # Check if all files belong to the same place and time 
            first = batch_meta[0]
            if any(m['station'] != first['station'] or m['month'] != first['month'] or m['year'] != first['year'] for m in batch_meta):
                st.error("🛑 **Batch Mismatch!** Ensure all files have the same Station, Month, and Year.")
                valid_batch = False

        if valid_batch:
            st.success(f"📂 **Validated Batch:** {first['station']} | {calendar.month_name[int(first['month'])]} {first['year']}")
            
            # Scan for Errors
            error_info = scan_files(files)
            
            if error_info:
                st.warning(f"⚠️ **Attention:** {len(error_info)} file(s) require correction.")
                for filename, rows in error_info.items():
                    st.write(f"❌ **{filename}**: Negative dips at rows `{rows}`")
                
                # The Action Button
                if st.button(f"🚀 Fix {len(error_info)} Files & Prepare Stage 2"):
                    clean_batch, fixed_only = process_batch_for_aggregation(files, error_info)
                    st.session_state['ready_batch'] = clean_batch
                    st.session_state['meta'] = batch_meta
                    
                    # ZIP DOWNLOAD CODE 
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for item in fixed_only:
                            zip_file.writestr(f"FIXED_{item['name']}", item['data'])
                    
                    st.success("✅ Errors fixed. Download the audit ZIP below to verify.")
                    st.download_button(
                        label="📥 Download Corrected Files (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name=f"FIXED_AUDIT_{first['station']}_{first['month']}.zip",
                        mime="application/zip"
                    )
            else:
                if st.button("🚀 All Clean - Proceed to Stage 2"):
                    clean_batch, _ = process_batch_for_aggregation(files, {})
                    st.session_state['ready_batch'] = clean_batch
                    st.session_state['meta'] = batch_meta
                    st.success("✅ Batch ready for aggregation.")


    # ==========================================
    # 3. STAGE 2: THE AUTOMATED ARCHITECT
    # ==========================================

    if st.session_state.get('ready_batch') is not None:
        st.markdown("---")
        st.header("Stage 2: Monthly Aggregation")
        
        meta = st.session_state['meta']
        station_val = meta[0]['station']
        year_val = meta[0]['year']
        month_abbr = calendar.month_name[int(meta[0]['month'])][:3].upper()
        
        # --- HMDAS-13 VALIDATION UPLOAD (XLS ONLY) ---
        st.subheader("🔍 HMDAS-13 Validation (Required)")
        st.info(f"Upload HMDAS-13 file matching: **13_HMDAS-13 6 hrly {station_val} {year_val}**")

        
        hmdas13_file = st.file_uploader(
            "Upload HMDAS-13 6hrly file",
            type=['xls'],
            accept_multiple_files=False,
            help="Filename must match: '13_HMDAS-13 6 hrly [STATION] [YYYY]' (.xls)"
        )
        
        # Validate HMDAS-13 filename
        hmdas13_valid = False
        if hmdas13_file:
            if not hmdas13_file.name.lower().endswith('.xls'):
                st.error("❌ **File type error!** Only .xls files accepted.")
            else:
                hmdas13_pattern = rf"13_HMDAS-13 6 hrly {re.escape(station_val)} {year_val}"
                if re.match(hmdas13_pattern, hmdas13_file.name):
                    hmdas13_valid = True
                    st.success(f"✅ **Valid HMDAS-13:** {hmdas13_file.name}")
                else:
                    st.error(f"❌ **Invalid filename!** Expected: `13_HMDAS-13 6 hrly {station_val} {year_val}.xls`")
                    st.info("Please rename your file to match the station and year from your batch.")
        else:
            st.warning("⚠️ Please upload the HMDAS-13 validation file (.xls) to proceed.")
        
        st.info("The cleaned data from Stage 1 is loaded. We will now populate the HMDAS-14 Master Template.")
        
        # Only show Build button if HMDAS-13 is valid
        if hmdas13_valid:
            st.session_state['h14_complete'] = True # Add this
            if st.button("🏗️ Build Master Report (HMDAS-14)"):
                with st.spinner("Executing Master Template Injection..."):
                    try:
                        HMDAS14_TEMPLATE_PATH = "templates/HMDAS-14 template.xlsm"
                        wb_master = openpyxl.load_workbook(HMDAS14_TEMPLATE_PATH, keep_vba=True)
                    except FileNotFoundError:
                        st.error("🚨 Template file not found! Please ensure 'HMDAS-14 Catarman 202110.xlsm' is in the app directory.")
                        st.stop()

                    # --- A. SPECIFIC CELL UPDATES (Sheet 1) --- 
                    ws1 = wb_master["1 Pattern 10min  Chart Reading"]
                    ws3 = wb_master["3 Corr Monthly RR Max"]
                    ws1['C2'] = month_abbr
                    ws3['N4'] = station_val
                    ws3['Q7'] = year_val

                    # --- B. EXTRACT FROM HMDAS-13 "data" SHEET ---
                    # Load HMDAS-13 file
                    try:
                        # Read "data" sheet from XLS file
                        df_data = pd.read_excel(
                            hmdas13_file, 
                            sheet_name="data",
                            engine='xlrd',  # Required for .xls files
                            header=None
                        )
                        
                        st.success("✅ HMDAS-13 'data' sheet loaded successfully!")
                        
                        # Find month in Row 8 (index 7, columns B=1, F=5, J=9, N=13, etc.)
                        month_row_idx = 7  # Row 8 (0-based)
                        month_full = calendar.month_name[int(meta[0]['month'])]
                        month_col_idx = None
                        
                        # Check columns B(1), F(5), J(9), N(13), etc. (every 4th column starting from B=1)
                        for col_idx in range(1, len(df_data.columns), 4):
                            cell_value = df_data.iloc[month_row_idx, col_idx]
                            if pd.notna(cell_value) and isinstance(cell_value, str) and month_full.lower() in str(cell_value).lower():
                                month_col_idx = col_idx
                                month_col_letter = openpyxl.utils.get_column_letter(month_col_idx + 1)  # +1 for Excel column
                                st.write(f"📊 Found {month_full} data in column {month_col_letter}")
                                break
                        
                        if month_col_idx is not None:
                            # Target sheet in master template
                            ws_pattern = wb_master["1 Pattern 10min  Chart Reading"]
                            
                            # Copy data: 31 rows x 4 columns (rows 10-40 → rows 3-33)
                            for row_offset in range(31):
                                src_row_idx = 9 + row_offset  # Row 10=idx9 (0-based)
                                tgt_row = 3 + row_offset
                                
                                # Copy 4 columns
                                for col_offset in range(4):
                                    src_col_idx = month_col_idx + col_offset
                                    tgt_col = 35 + col_offset  # AI=35, AJ=36, AK=37, AL=38
                                    
                                    if src_row_idx < len(df_data) and src_col_idx < len(df_data.columns):
                                        value = df_data.iloc[src_row_idx, src_col_idx]
                                        ws_pattern.cell(row=tgt_row, column=tgt_col).value = value
                            
                        else:
                            st.error(f"❌ {month_full} not found in Row 8 of 'data' sheet!")
                            
                    except Exception as e:
                        st.error(f"❌ Error reading HMDAS-13 file: {str(e)}")
                    
                    # --- C. DATA POPULATION (Monthly chart Reading) --- 
                    ws_monthly = wb_master["Monthly Rainfall Chart Reading"]
                    for i, item in enumerate(st.session_state['ready_batch']):
                        day_num = int(meta[i]['day'])
                        target_col = day_num + 1 
                            
                        wb_fixed = openpyxl.load_workbook(io.BytesIO(item['data']), data_only=True)
                        ws_fixed = wb_fixed.active
                            
                        for r in range(6, 151):
                            val = ws_fixed.cell(row=r, column=2).value
                            ws_monthly.cell(row=r, column=target_col).value = val



                    # --- D. EXECUTE CORRECTED MACRO LOGIC (CLEAN SHEET COPIES) ---
                  
                    ws1_master = wb_master["1 Pattern 10min  Chart Reading"]
                    ws2_template = wb_master["2 Data RR corrMAX"]
                    src_monthly = wb_master["Monthly Rainfall Chart Reading"]

                    valid_days = []
                    
                    # Iterate through columns 2-32 (Days 1-31)
                    for col in range(2, 33):
                        # 1. Check if day has data
                        day_data = [src_monthly.cell(row=r, column=col).value for r in range(6, 151)]
                        if all(v is None for v in day_data):
                            continue
                        
                        # 2. Extract Day Number
                        day_header = src_monthly.cell(row=5, column=col).value
                        try:
                            day_num = int(float(str(day_header).strip()))
                        except:
                            continue

                       # 3. Update the Master Input Sheet (Sheet 1)
                        # We populate AF3 and AE3 here so we can grab their values
                        # These correspond to columns 35 (AI) and 36 (AJ) in your HMDAS-13 extraction logic
                        for r_idx, val in enumerate(day_data):
                            ws1_master.cell(row=r_idx + 4, column=2).value = val
                        ws1_master['D2'].value = day_num

                        # 4. Create the New Sheet
                        new_sheet = wb_master.copy_worksheet(ws2_template)
                        new_sheet.title = f"{day_num:02d} {month_abbr} {year_val}"

                        # --- Updated Step 5: Manual B1 Calculation ---

                        # 1. Identify the Row in the AH3:AM33 table that matches the current day
                        target_row = None
                        for r in range(3, 34):
                            # Column 34 is AH (the Day column in your HMDAS-13 table)
                            if ws1_master.cell(row=r, column=34).value == day_num:
                                target_row = r
                                break

                        if target_row:
                            # 2. Calculate AF3 equivalent: Sum of AI:AL for this specific row
                            # If the cell value is "T" (case-insensitive) or None, it becomes 0
                            af3_sum = sum(
                                0 if str(ws1_master.cell(row=target_row, column=c).value).strip().upper() == 'T' 
                                else (ws1_master.cell(row=target_row, column=c).value or 0)
                                for c in range(35, 39) # AI=35 to AL=38
                            )
                            
                            # 3. Get B148 value from the current day's data
                            val_b148 = ws1_master.cell(row=148, column=2).value or 0
                            
                            # 4. Final Calculation: AF3 / B148
                            if isinstance(val_b148, (int, float)) and val_b148 != 0:
                                # Ensure af3_sum is a number before division
                                final_af3 = float(af3_sum) if not isinstance(af3_sum, str) else 0
                                calculated_b1 = final_af3 / val_b148
                            else:
                                calculated_b1 = 0
                        else:
                            calculated_b1 = 0

                        # 5. FREEZE THE RESULT
                        # Write the calculated number directly to B1, overwriting the formula string
                        new_sheet['B1'].value = calculated_b1

                        # Freeze the rainfall data in B4:B148 as before
                        for r_idx, val in enumerate(day_data):
                            new_sheet.cell(row=r_idx + 3, column=2).value = val

                        # 6. Pass this unique value to your HMDAS-15 report list
                        valid_days.append({
                            'name': new_sheet.title,
                            'day_num': day_num,
                            'worksheet': new_sheet,
                            'manual_b1': calculated_b1 
                        })
                    st.success(f"✅ {len(valid_days)} unique day-sheets created!")

                    # --- E. HMDAS-15: Validation (DISPLAY ON WEB WITH HIGHLIGHTING) ---
                    if valid_days:
                        if valid_days:st.session_state['h14_complete'] = True # Add this
                        st.subheader("📊 HMDAS-15 Validation Summary")
                        
                        # 1. Create a list of dictionaries
                        validation_rows = []
                        for day_info in valid_days:
                            b1_val = day_info['manual_b1']
                            is_pass = b1_val > 0.84
                            
                            validation_rows.append({
                                "Date/Sheet": day_info['name'],
                                "Correction Factor": b1_val,
                                "Status": "✅ PASS" if is_pass else "❌ REVIEW",
                                
                            })

                        # 2. Convert to Pandas DataFrame for styling
                        df = pd.DataFrame(validation_rows)

                        # 3. Define the styling function
                        def highlight_review(row):
                            # If status is REVIEW, color the whole row light red
                            if "REVIEW" in str(row.Status):
                                return ['background-color: #ffcccc'] * len(row)
                            return [''] * len(row)

                        # 4. Apply styling and display
                        styled_df = df.style.apply(highlight_review, axis=1)
                        
                        st.dataframe(styled_df, use_container_width=True, hide_index=True)

                        # --- F. Download HMDAS-14 ---
                        st.divider()
                        h14_buffer = io.BytesIO()
                        wb_master.save(h14_buffer)
                        h14_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download Updated HMDAS-14",
                            data=h14_buffer.getvalue(),
                            file_name=f"HMDAS-14_{station_val}_{month_abbr}_{year_val}.xlsm",
                            mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                        )

                    else:
                        st.error("No valid data found!")
        else:
            st.stop()



with tab2:
    st.header("📊 HMDAS-15 Annual Max Summary")

    if st.button("🧹 Reset Annual Process"):
        reset_app()

    st.info("Upload multiple HMDAS-14 files (different months) to compile the annual summary.")
    
    annual_files = st.file_uploader(
        "Upload HMDAS-14 Files (e.g., Catarman 202110.xlsm)", 
        type=["xlsm"], 
        accept_multiple_files=True,
        key=f"annual_up_{st.session_state['uploader_key']}"
    )
    
    from openpyxl.styles import PatternFill, Border, Side

    if annual_files:
        # --- VALIDATION LOGIC ---
        # Matches: HMDAS-14_[Station]_[Month]_[Year].xlsm
        name_pattern = r"HMDAS-14_(.+)_([A-Z]{3})_(\d{4})"
        batch_meta = []
        is_valid_batch = True

        for f in annual_files:
            match = re.search(name_pattern, f.name)
            if not match:
                st.error(f"❌ **Naming Error:** `{f.name}` does not follow the HMDAS-14 convention.")
                is_valid_batch = False
                break
            
            station, month, year = match.groups()
            batch_meta.append({'station': station.strip(), 'year': year, 'name': f.name})

        if is_valid_batch:
            first_file = batch_meta[0]
            mismatched_station = [m['name'] for m in batch_meta if m['station'] != first_file['station']]
            mismatched_year = [m['name'] for m in batch_meta if m['year'] != first_file['year']]

            if mismatched_station:
                st.error(f"🛑 **Station Mismatch!** Found different stations. All files must be from `{first_file['station']}`.")
                is_valid_batch = False
            elif mismatched_year:
                st.error(f"🛑 **Year Mismatch!** Found multiple years. All files must be from `{first_file['year']}`.")
                is_valid_batch = False

        if is_valid_batch:
            st.success(f"✅ **Validated Batch:** {batch_meta[0]['station']} | Year: {batch_meta[0]['year']}")
            st.session_state['h15_complete'] = True
            
            current_station = batch_meta[0]['station']
            current_year = batch_meta[0]['year']

            if st.button(f"🚀 Process & Generate HMDAS-15 {current_station}"):
                try:
                    extracted_data = []
                    wb_h15 = load_workbook(HMDAS15_TEMPLATE_PATH, keep_vba=True)
                    ws_summary = wb_h15.active
                    
                    # --- NEW UPDATES: Cells & Sheet Name ---
                    # 1. Update Sheet Title (Tab Name)
                    ws_summary.title = f"{current_year} {current_station} Max Summary"
                    
                    # 2. Update Cell A2
                    ws_summary['A2'].value = f"{current_station} Synop Station"
                    
                    # 3. Update Cell O18
                    ws_summary['O18'].value = f"{current_year} Annual Max per Duration"
                    
                    # --- STYLING ---
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    thick_border_side = Side(style='medium')
                    month_separator = Border(bottom=thick_border_side)
                    
                    lookback_steps = [1, 2, 3, 6, 12, 18, 36, 72, 144]

                    for uploaded_file in annual_files:
                        wb_h14 = load_workbook(uploaded_file, data_only=False) 
                        date_pattern = re.compile(r"^\d{2}\s[A-Z]{3}\s\d{4}$")
                        
                        for sheet_name in wb_h14.sheetnames:
                            if date_pattern.match(sheet_name):
                                ws_day = wb_h14[sheet_name]
                                b1_val = ws_day["B1"].value
                                multiplier = float(b1_val) if b1_val is not None else 1.0

                                col_c_data = []
                                for r in range(3, 148):
                                    val = ws_day.cell(row=r, column=2).value
                                    try:
                                        col_c_data.append((float(val) if val is not None else 0.0) * multiplier)
                                    except:
                                        col_c_data.append(0.0)

                                intensity_maxes = []
                                for lookback in lookback_steps:
                                    max_val = 0.0
                                    for i in range(lookback, len(col_c_data)):
                                        diff = col_c_data[i] - col_c_data[i - lookback]
                                        if diff > max_val:
                                            max_val = diff
                                    intensity_maxes.append(round(max_val, 1))

                                extracted_data.append({
                                    "date_obj": datetime.strptime(sheet_name, "%d %b %Y"),
                                    "label": sheet_name,
                                    "values": intensity_maxes
                                })

                    # Sort Chronologically
                    extracted_data.sort(key=lambda x: x['date_obj'])
                    
                    # 4. Injection Loop
                    last_month = None
                    for idx, entry in enumerate(extracted_data, 1):
                        target_row = 3 + idx
                        current_month = entry['date_obj'].month
                        
                        ws_summary.cell(row=target_row, column=1).value = idx
                        ws_summary.cell(row=target_row, column=2).value = entry['label']
                        ws_summary.cell(row=target_row, column=3).value = "Max"
                        
                        for c_offset, val in enumerate(entry['values']):
                            cell = ws_summary.cell(row=target_row, column=4 + c_offset)
                            cell.value = val
                            if val > 30:
                                cell.fill = red_fill
                        
                        # Apply monthly separator
                        if last_month is not None and current_month != last_month:
                            for col in range(1, 13):
                                ws_summary.cell(row=target_row - 1, column=col).border = month_separator
                        last_month = current_month

                    # 5. Save & Download
                    output_fn = f"HMDAS-15 {current_station} {current_year}.xlsm"
                    h15_buffer = io.BytesIO()
                    wb_h15.save(h15_buffer)
                    
                    st.success(f"✅ Processed successfully: {ws_summary.title}")
                    st.download_button(
                        label=f"📥 Download {output_fn}",
                        data=h15_buffer.getvalue(),
                        file_name=output_fn,
                        mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                    )

                except Exception as e:
                    st.error(f"Error during compilation: {e}")
